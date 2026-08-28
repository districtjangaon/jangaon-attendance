# -*- coding: utf-8 -*-
"""Build the daily attendance workbook locally, from the district summary files.

    python tools/daily-report-local.py [--day 27]

Produces docs/report/daily/Jangaon-attendance-<date>.xlsx with one row per
person on the rolls, marked or not. This is the same workbook the scheduled
email attaches; it exists so the report can be produced and circulated by hand
while the Apps Script send permission is being sorted out.

Two differences from the server-side version, both stated on the sheet:
  * Supervisors are absent. Their names live only in the live Users sheet;
    the field-staff export carries AWT and AWH.
  * The daily beneficiary return is recorded per centre, not per person, so
    the column reads for the centre the worker is posted to.
"""
import argparse
import csv
import io
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'docs', 'report', 'daily')


def load(*parts):
    with io.open(os.path.join(ROOT, 'summary', *parts), encoding='utf-8') as f:
        return json.load(f)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--day', default=None, help='day of month, e.g. 27')
    args = ap.parse_args()

    meta = load('meta.json')
    org = load('org.json')
    reports = load('reports', meta['month'] + '.json')

    # ------------------------------------------------------ establishment
    users = {}
    path = os.path.join(ROOT, 'master-data', 'IMPORT_USERS.csv')
    with io.open(path, encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            if r.get('user_id'):
                users[r['user_id']] = r

    sector_name = {s['code']: s['name'] for s in org['sectors']}
    awc_name = {k: v['n'] for k, v in org['awcs'].items()}

    # ------------------------------------------------------------- marks
    marks, leaves = {}, {}
    mdir = os.path.join(ROOT, 'summary', 'month')
    for fn in sorted(os.listdir(mdir)):
        if not fn.endswith('.json'):
            continue
        j = load('month', fn)
        for uid, days in (j.get('users') or {}).items():
            marks.setdefault(uid, {}).update(days)
        for uid, days in (j.get('leaves') or {}).items():
            leaves.setdefault(uid, {}).update(days)

    days_seen = sorted({d for v in marks.values() for d in v})
    day = args.day.zfill(2) if args.day else days_seen[-1]
    date_str = meta['month'] + '-' + day

    # The beneficiary return is filed per centre.
    reported_awcs = {a for a, v in (reports.get('awcs') or {}).items()
                     if day in (v.get('d') or {})}

    def fence(gf):
        return {'INSIDE': 'Yes', 'OUTSIDE': 'No'}.get(gf, 'No GPS fix')

    # The sheet is read by officers, not by the system that wrote the code.
    LEAVE_LABEL = {'CASUAL': 'Casual Leave', 'EARNED': 'Earned Leave',
                   'OPTIONAL': 'Optional Holiday', 'SICK': 'Medical Leave'}

    rows = []
    for uid, u in users.items():
        rec = (marks.get(uid) or {}).get(day) or {}
        i, o = rec.get('IN'), rec.get('OUT')
        leave = (leaves.get(uid) or {}).get(day, '')
        ever = bool(marks.get(uid))
        if leave:
            status = 'On leave'
        elif i and o:
            status = 'Present, day complete'
        elif i:
            status = 'Present, no OUT'
        elif o:
            status = 'OUT only'
        else:
            status = 'Not marked today' if ever else 'Never marked'
        rows.append([
            0, uid, u.get('name', ''), u.get('cadre', ''), u.get('project_code', ''),
            sector_name.get(u.get('sector_code', ''), u.get('sector_code', '')),
            awc_name.get(u.get('awc_id', ''), u.get('awc_id', '')),
            'Yes' if (i or o) else 'No',
            (i or {}).get('t', 'Not marked'), fence((i or {}).get('gf')) if i else '-',
            (i or {}).get('d', '') if i else '',
            (o or {}).get('t', 'Not marked'), fence((o or {}).get('gf')) if o else '-',
            (o or {}).get('d', '') if o else '',
            'Yes' if u.get('awc_id') in reported_awcs else 'No',
            LEAVE_LABEL.get(leave, leave) or 'No', 'Yes' if ever else 'No', status])

    rows.sort(key=lambda r: (r[5], r[6], r[2]))
    for n, r in enumerate(rows, 1):
        r[0] = n

    header = ['S.No', 'User ID', 'Name', 'Cadre', 'Project', 'Sector', 'Centre',
              'Attendance marked', 'Marked IN', 'IN inside geofence', 'IN distance (m)',
              'Marked OUT', 'OUT inside geofence', 'OUT distance (m)',
              'Daily report submitted', 'On leave', 'Ever marked', 'Status']

    # ------------------------------------------------------------- write
    os.makedirs(OUT, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance ' + date_str

    ws.append(header)
    head_fill = PatternFill('solid', fgColor='0B5C4F')
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = head_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    for r in rows:
        ws.append(r)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(header)), len(rows) + 1)
    widths = [6, 10, 30, 7, 8, 18, 26, 11, 11, 12, 12, 11, 12, 12, 13, 12, 9, 22]
    for n, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(n)].width = w
    ws.row_dimensions[1].height = 30

    # A short second sheet, so nobody has to reconstruct the totals by hand.
    s2 = wb.create_sheet('Summary')
    marked = sum(1 for r in rows if r[7] == 'Yes')
    on_leave = sum(1 for r in rows if r[15] != 'No')
    never = sum(1 for r in rows if r[16] == 'No')
    for line in [
        ['Jangaon District - daily attendance', ''],
        ['Date', date_str],
        ['Source', 'summary files generated ' + meta['generatedAt']],
        ['', ''],
        ['On the rolls (AWT and AWH)', len(rows)],
        ['Attendance marked', marked],
        ['Marked IN', sum(1 for r in rows if r[8] != 'Not marked')],
        ['Marked OUT', sum(1 for r in rows if r[11] != 'Not marked')],
        ['IN inside the centre boundary', sum(1 for r in rows if r[9] == 'Yes')],
        ['IN outside the centre boundary', sum(1 for r in rows if r[9] == 'No')],
        ['IN with no usable GPS fix', sum(1 for r in rows if r[9] == 'No GPS fix')],
        ['On sanctioned leave', on_leave],
        ['Not marked and not on leave', sum(1 for r in rows if r[7] == 'No' and r[15] == 'No')],
        ['Never marked since deployment', never],
        ['Centres filing the daily return', len(reported_awcs)],
        ['', ''],
        ['Not included', 'Supervisors: their names are held only in the live Users '
                         'sheet, not in the field-staff export this was built from.'],
        ['Note', 'The daily beneficiary return is filed per centre, so that column '
                 'reads for the centre the worker is posted to.'],
    ]:
        s2.append(line)
    s2.column_dimensions['A'].width = 34
    s2.column_dimensions['B'].width = 80
    for r in range(1, s2.max_row + 1):
        s2.cell(row=r, column=1).font = Font(bold=True)
        s2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    out = os.path.join(OUT, 'Jangaon-attendance-%s.xlsx' % date_str)
    wb.save(out)

    print('Workbook: %s' % out)
    print('  %d persons, %d marked, %d not marked and not on leave, %d never marked'
          % (len(rows), marked, sum(1 for r in rows if r[7] == 'No' and r[15] == 'No'), never))


if __name__ == '__main__':
    main()
