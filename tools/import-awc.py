#!/usr/bin/env python3
"""
import-awc.py — converts the department AWC master xlsx into the five
IMPORT_* CSVs consumed by backend importFromSheets(), plus a data-issues
report for the department to fix.

Input  : input/AWC DATA-31.07.2026.xlsx  (Sl.No, Project, Sector, AWC Name,
         AWT Name, AWT Mobile, AWH Name, AWH Mobile, Latitude, Longitude;
         2 header rows)
Output : master-data/IMPORT_PROJECTS.csv   project_code,name
         master-data/IMPORT_SECTORS.csv    sector_code,project_code,name,supervisor_phone
         master-data/IMPORT_AWCS.csv       awc_id,sector_code,project_code,name,lat,lng,radius_m
         master-data/IMPORT_USERS.csv      user_id,phone,name,cadre,role,project_code,sector_code,awc_id
         master-data/IMPORT_SCHEDULES.csv  project_code,cadre,in_start,in_end,late_after,out_start,out_end
         master-data/ISSUES.md             everything the department must fix

master-data/ holds REAL personal data — it is gitignored and must never be
committed or shared outside the department.

Usage: python tools/import-awc.py [path-to-xlsx]
Requires: pip install openpyxl
"""
import csv
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / 'input' / 'AWC DATA-31.07.2026.xlsx'
OUT = ROOT / 'master-data'

# District bounding box: coords outside it are treated as data-entry errors,
# blanked (marks become UNVERIFIED, not falsely OUTSIDE) and reported.
LAT_MIN, LAT_MAX = 17.2, 18.2
LNG_MIN, LNG_MAX = 78.7, 79.7

PROJECT_CODES = {'JANGAON': 'JGN', 'KODAKANDLA': 'KDK', 'STN.GHANPUR': 'SGN'}
DEFAULT_RADIUS_M = 200

# Default AWC working hours (editable later via IMPORT_SCHEDULES re-import).
SCHEDULE_ROWS = [
    ('ALL', 'ALL', '08:30', '10:30', '09:30', '15:30', '17:30'),
]

VACANT = {'vacant', '', 'none', '-', 'nil', 'na'}


def is_vacant(v):
    return v is None or str(v).strip().lower() in VACANT


def clean_name(v):
    s = re.sub(r'\s+', ' ', str(v).strip())
    # Title-case fully-upper names; leave mixed-case names as entered.
    return s.title() if s.isupper() else s


def norm_phone(v):
    if v is None:
        return None
    s = re.sub(r'\D', '', str(v))
    if len(s) == 12 and s.startswith('91'):
        s = s[2:]
    return s if len(s) == 10 else None


def sector_title(s):
    s = re.sub(r'\s+', ' ', str(s).strip())
    return s.title() if (s.isupper() or s.islower()) else s


def main():
    if not XLSX.exists():
        sys.exit(f'FATAL: input file not found: {XLSX}')
    OUT.mkdir(exist_ok=True)

    ws = openpyxl.load_workbook(XLSX, data_only=True)['Sheet1']
    header = [str(c or '').strip() for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    expect = ['Sl.No', 'Project', 'Sector', 'AWC Name', 'AWT Name', 'AWT Mobile Number',
              'AWH NAME', 'AWH Mobile Number', 'Latitude', 'Longitude']
    if header != expect:
        sys.exit(f'FATAL: header row 2 changed.\n  expected: {expect}\n  found:    {header}')

    rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
            if any(c is not None and str(c).strip() for c in r)]

    issues = {
        'bad_coords': [], 'dup_awc': [], 'bad_phone': [], 'cross_awc_phone': [],
        'awt_vacant': [], 'awh_vacant': [], 'both_vacant': [],
    }

    projects = {}        # code -> name
    sectors = {}         # (proj_code, name) -> sector_code
    awcs = []            # dicts
    users = []           # dicts
    seen_awc_names = {}  # (proj, sector, name) -> count

    for r in rows:
        proj_name = str(r[1]).strip().upper()
        if proj_name not in PROJECT_CODES:
            sys.exit(f'FATAL: unknown project {proj_name!r} at Sl.No {r[0]} — add it to PROJECT_CODES.')
        pcode = PROJECT_CODES[proj_name]
        projects[pcode] = proj_name.title().replace('Stn.', 'Stn. ')

        sname = sector_title(r[2])
        skey = (pcode, sname)
        if skey not in sectors:
            sectors[skey] = f'S{len(sectors) + 1:02d}'
        scode = sectors[skey]

        name = re.sub(r'\s+', ' ', str(r[3]).strip())
        nkey = (pcode, sname, name.lower())
        n = seen_awc_names.get(nkey, 0) + 1
        seen_awc_names[nkey] = n
        if n > 1:
            issues['dup_awc'].append(f'Sl.No {r[0]}: "{name}" appears {n}x in {proj_name}/{sname} — renamed "{name} ({n})"')
            name = f'{name} ({n})'

        lat, lng = r[8], r[9]
        try:
            lat, lng = round(float(lat), 6), round(float(lng), 6)
            if not (LAT_MIN <= lat <= LAT_MAX and LNG_MIN <= lng <= LNG_MAX):
                issues['bad_coords'].append(
                    f'Sl.No {r[0]} {name}: ({lat}, {lng}) outside district box — blanked, re-capture on site')
                lat = lng = ''
        except (TypeError, ValueError):
            issues['bad_coords'].append(f'Sl.No {r[0]} {name}: missing/non-numeric coordinates — blanked')
            lat = lng = ''

        awc_id = f'A{len(awcs) + 1:04d}'
        awcs.append({'awc_id': awc_id, 'sector_code': scode, 'project_code': pcode,
                     'name': name, 'lat': lat, 'lng': lng, 'radius_m': DEFAULT_RADIUS_M})

        awt_vac, awh_vac = is_vacant(r[4]), is_vacant(r[6])
        if awt_vac and awh_vac:
            issues['both_vacant'].append(f'{awc_id} {name}')
        elif awt_vac:
            issues['awt_vacant'].append(f'{awc_id} {name}')
        elif awh_vac:
            issues['awh_vacant'].append(f'{awc_id} {name}')

        for name_col, phone_col, cadre in ((4, 5, 'AWT'), (6, 7, 'AWH')):
            if is_vacant(r[name_col]):
                continue
            phone = norm_phone(r[phone_col])
            if phone is None:
                issues['bad_phone'].append(
                    f'Sl.No {r[0]} {name} {cadre} {clean_name(r[name_col])}: mobile {r[phone_col]!r} invalid — '
                    'imported without phone, CANNOT LOGIN until fixed')
                phone = ''
            users.append({'user_id': f'U{len(users) + 1:04d}', 'phone': phone,
                          'name': clean_name(r[name_col]), 'cadre': cadre, 'role': 'FIELD',
                          'project_code': pcode, 'sector_code': scode, 'awc_id': awc_id})

    # Phones shared across DIFFERENT AWCs (same-AWC AWT/AWH sharing the centre
    # phone is normal and handled by the app's user picker).
    by_phone = {}
    for u in users:
        if u['phone']:
            by_phone.setdefault(u['phone'], []).append(u)
    for phone, us in sorted(by_phone.items()):
        if len({u['awc_id'] for u in us}) > 1:
            issues['cross_awc_phone'].append(
                phone + ': ' + '; '.join(f"{u['cadre']} {u['name']} @ {u['awc_id']}" for u in us))

    def write_csv(fname, fieldnames, dicts):
        with open(OUT / fname, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(dicts)

    write_csv('IMPORT_PROJECTS.csv', ['project_code', 'name'],
              [{'project_code': c, 'name': n} for c, n in sorted(projects.items())])
    write_csv('IMPORT_SECTORS.csv', ['sector_code', 'project_code', 'name', 'supervisor_phone'],
              sorted(({'sector_code': c, 'project_code': p, 'name': n, 'supervisor_phone': ''}
                      for (p, n), c in sectors.items()), key=lambda d: d['sector_code']))
    write_csv('IMPORT_AWCS.csv', ['awc_id', 'sector_code', 'project_code', 'name', 'lat', 'lng', 'radius_m'], awcs)
    write_csv('IMPORT_USERS.csv', ['user_id', 'phone', 'name', 'cadre', 'role',
                                   'project_code', 'sector_code', 'awc_id'], users)
    write_csv('IMPORT_SCHEDULES.csv', ['project_code', 'cadre', 'in_start', 'in_end',
                                       'late_after', 'out_start', 'out_end'],
              [dict(zip(['project_code', 'cadre', 'in_start', 'in_end', 'late_after',
                         'out_start', 'out_end'], row)) for row in SCHEDULE_ROWS])

    # Optional: state-government holiday list (Sundays are computed by rule,
    # they do not belong in this file). Columns: Date | Occasion.
    hol_xlsx = ROOT / 'input' / 'holidays.xlsx'
    if hol_xlsx.exists():
        hws = openpyxl.load_workbook(hol_xlsx, data_only=True).worksheets[0]
        hol_rows = []
        for r in hws.iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            d = r[0].strftime('%Y-%m-%d') if hasattr(r[0], 'strftime') else str(r[0]).strip()
            hol_rows.append({'date': d, 'name': str(r[1] or 'Holiday').strip()})
        write_csv('IMPORT_HOLIDAYS.csv', ['date', 'name'], hol_rows)
        print(f'holidays: {len(hol_rows)} from {hol_xlsx.name}')

    with open(OUT / 'ISSUES.md', 'w', encoding='utf-8') as f:
        f.write(f'# Data issues — {XLSX.name}\n\n')
        f.write(f'Imported: {len(awcs)} AWCs, {len(sectors)} sectors, {len(projects)} projects, '
                f'{len(users)} staff ({sum(1 for u in users if u["cadre"] == "AWT")} AWT, '
                f'{sum(1 for u in users if u["cadre"] == "AWH")} AWH).\n\n')
        f.write('Staffing model: every person in the sheet is a plain field user. There are '
                'no sector supervisors or CDPOs. Add the console accounts as rows in '
                'IMPORT_USERS.csv before importing (ids from U2001): the technical '
                'admin and the Collector, each with cadre OTHER and role ADMIN. '
                'Leave `supervisor_phone` in IMPORT_SECTORS.csv empty.\n\n')
        titles = {
            'bad_coords': 'Coordinates outside the district (blanked — geofence UNVERIFIED until re-captured on site)',
            'dup_awc': 'Duplicate AWC names (auto-renamed)',
            'bad_phone': 'Invalid mobile numbers (user cannot login until corrected)',
            'cross_awc_phone': 'Same mobile used at more than one AWC (verify with the sectors)',
            'awt_vacant': 'AWCs with AWT post vacant',
            'awh_vacant': 'AWCs with AWH post vacant',
            'both_vacant': 'AWCs with BOTH posts vacant (no one can mark attendance there)',
        }
        for key, title in titles.items():
            f.write(f'## {title} — {len(issues[key])}\n\n')
            for line in issues[key]:
                f.write(f'- {line}\n')
            f.write('\n')

    print(f'AWCs: {len(awcs)}  sectors: {len(sectors)}  projects: {len(projects)}  users: {len(users)}')
    for k, v in issues.items():
        print(f'  {k}: {len(v)}')
    print(f'Wrote {OUT}')


if __name__ == '__main__':
    main()
